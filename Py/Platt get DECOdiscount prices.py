#-------------------------------------------------------------------------------------------------
#   Required Installations
#   Install BS4 package
#   Install Requests Package
#   Install Selenium
#-------------------------------------------------------------------------------------------------

import requests
import time
#Import Beautiful Soup for web scraping
from bs4 import BeautifulSoup
import pathlib
# Import openpyxl for excel workbook control
import openpyxl
from openpyxl import Workbook

# Import Selenium driver for web scraping and error handling during webscrape
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.common.exceptions import ElementNotInteractableException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
options = Options()
options.headless = True
driver = webdriver.Chrome(chrome_options=options)


from random import randrange                
import datetime

#Import text analysis tool
import re


# Set filename 
filename = "Commodity Price Table"

workbook_loc = str(pathlib.Path(__file__).parent.parent.absolute()) + "\Commodity_Price_Table.xlsx" #Universal

wb = openpyxl.load_workbook(workbook_loc)

# grab the commodity table worksheet
worksheet = wb["Commodity Table"]

# iterate over the cells in the row to get excel sheet column positions 
for cell in worksheet[1]: #only scan first row
    # check if the cell's value matches the search value
    if cell.value == 'Platt_discount_ID':
        print("ID column->" + str(cell.column))
        ID_Column = cell.column
    if cell.value == 'Platt_discount_Product':
        print("Description column->" + str(cell.column))
        Description_Column = cell.column
    if cell.value == 'Platt_discount_UPC':
        print("Platt_discount UPC column->" + str(cell.column))
        Product_UPC_Column = cell.column
    if cell.value == 'Platt_discount_Product':
        print("Platt_discount Product name column->" + str(cell.column))
        Product_Column = cell.column
    if cell.value == 'Platt_discount_Cost':
        print("Cost column->" + str(cell.column))
        Cost_column = cell.column
    if cell.value == 'Platt_discount_Unit':
        print("Unit column->" + str(cell.column))
        Unit_Column = cell.column
    if cell.value == 'Platt_discount_Src':
        print("Src column->" + str(cell.column))
        Src_Column = cell.column
    if cell.value == 'Platt_discount_Update':
        print("Update column->" + str(cell.column))
        Update_Column = cell.column
    #Last column to search for before breaking
        break

driver = webdriver.Chrome()

# Log into Platt_discount
driver.get("https://auth.rexelusa.com/login?returnUrl=%2Fconnect%2Fauthorize%2Fcallback%3Fprotocol%3Doauth2%26response_type%3Dcode%26access_type%3Doffline%26client_id%3Dstorefront-web-v2%26redirect_uri%3Dhttps%253A%252F%252Fwww.Platt_discount.com%252Fcallback%26scope%3Dsf.web%2520offline_access%26state%3DCqO1Txt-2dAGkPKn9K6jq%26code_challenge_method%3DS256%26banner%3DPlatt_discount%26code_challenge%3D45CX5YBoL1KsZlN4oZXiH4nE-rAeQgjHSBLOnuE7uwI")
driver.implicitly_wait (10)

# enter the login details
username_field = driver.find_element_by_xpath('//*[@id="input-19"]')
username_field.send_keys('Abbott@Duttonelectric.com')


# submit the login form
submit_button = driver.find_element_by_xpath('//*[@id="app"]/div/main/div/div/div[1]/div[1]/div[3]/button/span')
submit_button.click()
print (" ")
print ("Logging into PLatt....")
print (" ")
time.sleep(2)

password_field = driver.find_element_by_xpath('//*[@id="input-19"]')
password_field.send_keys('J@bbott16')

# submit the login form
submit_button = driver.find_element_by_xpath('//*[@id="app"]/div/main/div/div/div[1]/div[1]/div[3]/button/span')
submit_button.click()
password_field.send_keys('~')
time.sleep(5)

# Reload PLatt page
driver.get("https://www.platt.com/login")
driver.implicitly_wait (10)


example_class_elements = driver.find_elements_by_class_name("v-list-item__content")
# Print the inner text of each element
#for element in example_class_elements:
#    print(element.text)


print (" ")
print ("Logged into Platt_discount")
print (" ")

# begin main loop getting data
for x in range(2, 580): #start no lower than 2, since 1 is the column header
    #create save ticker for save intervals
    save_counter = 0
    #Create info ticker for user messaging 
    New_CPB_info = 0

    # Get Source webURLlink to product data
    Src = worksheet.cell(row=x, column=Src_Column).value

    # Only continue code if there is a Src
    if Src:

        # Get Commodity Sheet Description
        ItemDescription = worksheet.cell(row=x, column=Description_Column).value
        # Get Commodity Sheet Description VendorProductID (Per Vendor)
        VendorProductID = worksheet.cell(row=x, column=ID_Column).value
        # ProductUPC Commodity Sheet Description (Per manufacturer of item)
        ProductUPC = worksheet.cell(row=x, column=Product_UPC_Column).value
        # Product Name
        ProductName = worksheet.cell(row=x, column =Product_Column).value

        # Check if a url link to product was on the Commodity sheet, or search
        if Src:
            # A url link to product is already on the Db
            url = Src

            driver.get(Src)

            # Get 'price' vi beautifulsoup
            Item_Cost = driver.find_element_by_class_name ('text--black').text

            # Write Price to commodities Table
            #print (Item_Cost)    
            worksheet.cell(row=x, column=Cost_column, value=Item_Cost)
            
            
            # Write the updated time to Comodity Sheet
            worksheet.cell(row=x, column=Update_Column, value= datetime.datetime.now())

            #Give User Feedback on what scraped
            print("Platt_discount   row:" + str(x) + " Item:" + str(ProductName)  + "                  $" + str(Item_Cost))
            if New_CPB_info  == 1:
                print("Platt_discount Tree-> Electrical->" + str(CPB02) + "->" + str(CPB03) + "->" + str(CPB04) + "->" + str(CPB05))

            #Make progress saves at intervals
            save_counter = save_counter + 1
            if save_counter > 100:
                wb.save(filename=workbook_loc)
                save_counter = 0
                print ("Saved")

#Save when complete
wb.save(filename=workbook_loc)
print ("Complete & Saved")
