




#-------------------------------------------------------------------------------------------------
#   Required Installations
#   Install BS4 package
#   Install Requests Package
#   Install Selenium
#-------------------------------------------------------------------------------------------------

import requests
import time
#Import Beautiful Soup for web scraping
from bs4 import BeautifulSoup
import pathlib
# Import openpyxl for excel workbook control
import openpyxl
from openpyxl import Workbook

# Import Selenium driver for web scraping and error handling during webscrape
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.common.exceptions import ElementNotInteractableException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException


options = Options()
options.headless = True
#driver = webdriver.Chrome(chrome_options=options)
driver = webdriver.Chrome()

from random import randrange                
import datetime

#Import text analysis tool
import re


# Load Workbook
filename = "Commodity Price Table"
#           Commodity Price Table
workbook_loc = str(pathlib.Path(__file__).parent.parent.absolute()) + "\Commodity_Price_Table.xlsx" #Universal
wb = openpyxl.load_workbook(workbook_loc)

# grab the commodity table worksheet
worksheet = wb["Commodity Table"]

# iterate over the cells in the row to get excel sheet column positions 
for cell in worksheet[1]: #only scan first row
    # check if the cell's value matches the search value
        #Last column to search for before breaking

    if cell.value == 'Graybar_discount_ID':
        print("ID column->" + str(cell.column))
        Graybar_discount_ID_Column = cell.column
    if cell.value == 'Graybar_discount_UPC':
        print("Graybar_discount UPC column->" + str(cell.column))
        Product_UPC_Column = cell.column
    if cell.value == 'Graybar_discount_Product':
        print("Graybar_discount Product name column->" + str(cell.column))
        Product_Column = cell.column
    if cell.value == 'Graybar_discount_Cost':
        print("Cost column->" + str(cell.column))
        Cost_column = cell.column
    if cell.value == 'Graybar_discount_Unit': #Ref->Graybar_discount
        print("Unit column->" + str(cell.column))
        Unit_Column = cell.column
    if cell.value == 'Graybar_discount_Src':
        print("Src column->" + str(cell.column))
        Src_Column = cell.column
    if cell.value == 'Graybar_discount_Update':
        print("Update column->" + str(cell.column))
        Update_Column = cell.column
        break


#driver = webdriver.Chrome()
#driver = webdriver.Chrome(chrome_options=options)


# Load Graybar Log In Page
driver.get('https://www.graybar.com/login')
driver.implicitly_wait(60)




# Find the input element using its name attribute
input_element = driver.find_element('xpath','//*[@id="j_username"]')

# Enter new data into the input field
input_element.send_keys("Abbott@Duttonelectric.com")

# Find the input element using its name attribute
input_element = driver.find_element('xpath','//*[@id="j_password"]')

# Enter new text into the input field
input_element.send_keys("!Mongoose2023")

# Find the button element using its data-qa attribute
button = driver.find_element('xpath','//*[@id="loginButton"]')


# Click the button
button.click()
#button.send_keys ("~")
driver.implicitly_wait (30)
#wait for pop up
time.sleep(10)

#get past pop up  'we welcome your feedback screen'
try:
    element = driver.find_element ("xpath", '//*[@id="fsrInvite"]/section[3]/button[2]')
    element.click () 
except NoSuchElementException:
    print ("no survery screen, awesome, let's continue")
    
#create save ticker for save intervals
save_counter = 0

# begin main loop getting data
for x in range(2, 1000): #start no lower than 2, since 1 is the column header

    # Get General Product Description 
    product_description = worksheet.cell(row=x, column=1).value

    # Get Source webURLlink to product data
    Src = worksheet.cell(row=x, column=Src_Column).value
  
          
    # Check if a url link to product was on the Commodity sheet, or search
    if Src:
        if "graybar" in Src:
            if "content" not in Src:
                # A url link to product is already on the Db
                driver.get(Src)    
                driver.implicitly_wait(10)


                # Create a BeautifulSoup object
                html = driver.page_source
                soup = BeautifulSoup(html, "html.parser")

                try:
                    # Get 'price' vi beautifulsoup
                    price_element = soup.find('p', {'class': 'price'})
                    priceperqty = price_element.text.strip()
                    
                    priceperqty = priceperqty.replace("$","")
                    priceperqty = priceperqty.replace("/ ft","")
                    priceperqty = priceperqty.replace("/ FT","")
                    priceperqty = priceperqty.replace("/ ea","")
                    priceperqty = priceperqty.replace("/ EA","")
                    
                    uom = 1

                    if "/ 1000" in priceperqty:
                        priceperqty = priceperqty.replace("/ 1000FT","") #/ 1000ft
                        priceperqty = priceperqty.replace("/ 1000ft","")
                        priceperqty = priceperqty.replace("/ 1000EA","")
                        priceperqty = priceperqty.replace("/ 1000ea","")
                        priceperqty = priceperqty.replace(",","")
                        uom = 1000

                    if "/ 100" in priceperqty:
                        priceperqty = priceperqty.replace("/ 100FT","")
                        priceperqty = priceperqty.replace("/ 100ft","")
                        priceperqty = priceperqty.replace("/ 100EA","")
                        priceperqty = priceperqty.replace("/ 100ea","")
                        priceperqty = priceperqty.replace(",","")
                        uom = 100

                    priceperqty = float(priceperqty)
                    priceperqty = priceperqty/uom

                except AttributeError:
                    priceperqty = 0

                # Get Graybar ID (also UPC?)
                if priceperqty != 0:
                    # find the span element by class name
                    span_elem = soup.find_all('span', {'class': 'code'})[1]
                    # get the text content of the span element
                    Graybar_discount_ID = span_elem.text
                    #Graybar_discount_ID_Column
                    if worksheet.cell(row=x, column=Graybar_discount_ID_Column).value != Graybar_discount_ID:
                        worksheet.cell(row=x, column=Graybar_discount_ID_Column, value = Graybar_discount_ID)

                # write price to the worksheet
                worksheet.cell(row=x, column=Cost_column, value=priceperqty)

                # Write the updated time to Comodity Sheet
                worksheet.cell(row=x, column=Update_Column, value= datetime.datetime.now())


                #Give User Feedback on what scraped
                print("Graybar_discount   row:" + str(x) + " " + str(product_description) +     " $" + str(priceperqty) + " ")
        

                #Make progress saves at intervals
                save_counter = save_counter + 1
                if save_counter > 100:
                    wb.save(filename=workbook_loc)
                    save_counter = 0
                    print ("Saved")

#Save when complete
wb.save(filename=workbook_loc)
print ("Complete & Saved")
