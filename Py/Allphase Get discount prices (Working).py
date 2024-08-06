




#-------------------------------------------------------------------------------------------------
#   Required Installations
#   Install BS4 package
#   Install Requests Package
#   Install Selenium
#-------------------------------------------------------------------------------------------------

import requests
import time
#Import Beautiful Soup for web scraping
from bs4 import BeautifulSoup
import pathlib
# Import openpyxl for excel workbook control
import openpyxl
from openpyxl import Workbook

# Import Selenium driver for web scraping and error handling during webscrape
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.common.exceptions import ElementNotInteractableException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By

options = Options()
options.headless = True

from random import randrange                
import datetime

#Import text analysis tool
import re


# Load Workbook
filename = "Commodity Price Table"
#           Commodity Price Table
workbook_loc = str(pathlib.Path(__file__).parent.parent.absolute()) + "\Commodity_Price_Table.xlsx" #Universal
wb = openpyxl.load_workbook(workbook_loc)

# grab the commodity table worksheet
worksheet = wb["Commodity Table"]

# iterate over the cells in the row to get excel sheet column positions 
for cell in worksheet[1]: #only scan first row
    # check if the cell's value matches the search value
        #Last column to search for before breaking

    if cell.value == 'Wesco_discount_ID':
        print("ID column->" + str(cell.column))
        Wesco_discount_ID_Column = cell.column
    if cell.value == 'Wesco_discount_UPC':
        print("Wesco_discount UPC column->" + str(cell.column))
        Product_UPC_Column = cell.column
    if cell.value == 'Wesco_discount_Product':
        print("Wesco_discount Product name column->" + str(cell.column))
        Product_Column = cell.column
    if cell.value == 'Wesco_discount_Cost':
        print("Cost column->" + str(cell.column))
        Cost_column = cell.column
    if cell.value == 'Wesco_discount_Unit': #Ref->Wesco_discount
        print("Unit column->" + str(cell.column))
        Unit_Column = cell.column
    if cell.value == 'Wesco_discount_Src':
        print("Src column->" + str(cell.column))
        Src_Column = cell.column
    if cell.value == 'Wesco_discount_Update':
        print("Update column->" + str(cell.column))
        Update_Column = cell.column
        break


driver = webdriver.Chrome()
#driver = webdriver.Chrome(chrome_options=options)


# Load Wesco Log In Page
driver.get('https://buy.wesco.com/login')
driver.implicitly_wait(10)




# Find the input element using its name attribute
input_element = driver.find_element ('xpath','/html/body/div[3]/div[4]/div[2]/div[1]/form/input[1]')

# Enter new data into the input field
input_element.send_keys("Abbott@Duttonelectric.com")

# Find the input element using its name attribute
input_element = driver.find_element ('xpath', '/html/body/div[3]/div[4]/div[2]/div[1]/form/input[2]')

# Enter new text into the input field
input_element.send_keys("J@bbott16")

#click on cookies button to make it go away
input_element = driver.find_element ('xpath', '//*[@id="hs-eu-confirmation-button"]')
input_element.click ()
driver.implicitly_wait(10)

# Find the button element using its data-qa attribute
button = driver.find_element ('xpath','/html/body/div[3]/div[4]/div[2]/div[1]/form/div/button')

# Click the button
button.click()
#button.send_keys ("~")

#driver.implicitly_wait (30)
#wait for user to navigate CAPTCHA
#time.sleep(30)

# begin main loop getting data
for x in range(2, 100): #start no lower than 2, since 1 is the column header
    #create save ticker for save intervals
    save_counter = 0

    # Get General Product Description 
    product_description = worksheet.cell(row=x, column=1).value

    # Get Source webURLlink to product data
    Src = worksheet.cell(row=x, column=Src_Column).value
  
          
    # Check if a url link to product was on the Commodity sheet, or search
    if Src:
        if "wesco" in Src:
            if "content" not in Src:
                # A url link to product is already on the Db
                driver.get(Src)    
                driver.implicitly_wait(10)

                # Create a BeautifulSoup object
                html = driver.page_source
                soup = BeautifulSoup(html, "html.parser")

                #time.sleep (1000)
                
                try:
                    # Get 'price'
                    # Find the <span> tag with class "priceValue"
                    span_tag = driver.find_element('xpath', '/html/body/div[2]/div[4]/div[2]/div/div[4]/div[2]/div/div[2]/div[1]/span')

                    # Get the text of the <span> tag
                    priceperqty = span_tag.text
                    priceperqty = priceperqty.replace("$","")
                    #print (priceperqty)
                except NoSuchElementException:
                    print ("No price on page")
                    priceperqty = "0"

                #uom = driver.find_element_by_class_name("price_per").text
                #if "each" not in uom:
                #    time.sleep(0.1)
                #print (uom)

                sku_element = soup.find('h1', class_='c-product-detail-base__name')

                # write price to the worksheet
                worksheet.cell(row=x, column=Cost_column, value=priceperqty)

                # Write the updated time to Comodity Sheet
                worksheet.cell(row=x, column=Update_Column, value= datetime.datetime.now())


                #Give User Feedback on what scraped
                print("Wesco_discount   row:" + str(x) + " " + str(product_description) +     " $" + str(priceperqty) + " ")
        

                #Make progress saves at intervals
                save_counter = save_counter + 1
                if save_counter > 100:
                    wb.save(filename=workbook_loc)
                    save_counter = 0
                    print ("Saved")

#Save when complete
wb.save(filename=workbook_loc)
print ("Complete & Saved")
