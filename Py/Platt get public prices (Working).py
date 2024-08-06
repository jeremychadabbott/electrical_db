





#-------------------------------------------------------------------------------------------------
#   Required Installations
#   Install BS4 package
#   Install Requests Package
#   Install Selenium
#-------------------------------------------------------------------------------------------------


import requests
import time
#Import Beautiful Soup for web scraping
from bs4 import BeautifulSoup
import pathlib
# Import openpyxl for excel workbook control
import openpyxl
from openpyxl import Workbook

# Import Selenium driver for web scraping and error handling during webscrape
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.common.exceptions import ElementNotInteractableException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
options = Options()
options.headless = True
driver = webdriver.Chrome(chrome_options=options)


from random import randrange                
import datetime

#Import text analysis tool
import re


# Load Workbook
filename = "Commodity Price Table"
#           Commodity Price Table
workbook_loc = str(pathlib.Path(__file__).parent.parent.absolute()) + "\Commodity_Price_Table.xlsx" #Universal
wb = openpyxl.load_workbook(workbook_loc)

# grab the commodity table worksheet
worksheet = wb["Commodity Table"]

# iterate over the cells in the row to get excel sheet column positions 
for cell in worksheet[1]: #only scan first row
    # check if the cell's value matches the search value
    if cell.value == 'UPC':
        print("UPC column->" + str(cell.column))
        UPC_Column = cell.column
    if cell.value == 'Platt_ID':
        print("ID column->" + str(cell.column))
        ID_Column = cell.column
    if cell.value == 'Platt_Product':
        print("Description column->" + str(cell.column))
        Description_Column = cell.column
    if cell.value == 'Platt_UPC':
        print("Platt UPC column->" + str(cell.column))
        Product_UPC_Column = cell.column
    if cell.value == 'Platt_Product':
        print("Platt Product name column->" + str(cell.column))
        Product_Column = cell.column
    if cell.value == 'Platt_Cost':
        print("Cost column->" + str(cell.column))
        Cost_column = cell.column
    if cell.value == 'Platt_Unit':
        print("Unit column->" + str(cell.column))
        Unit_Column = cell.column
    if cell.value == 'Platt_Src':
        print("Src column->" + str(cell.column))
        Src_Column = cell.column
    if cell.value == 'Platt_Update':
        print("Update column->" + str(cell.column))
        Update_Column = cell.column

    #Get Common Part Branch (CPB) Columns
    if cell.value == 'Platt CPB01':
        print("Platt CPB01->" + str(cell.column))
        CPB01_Column = cell.column
    if cell.value == 'Platt CPB02':
        print("Platt CPB02->" + str(cell.column))
        CPB02_Column = cell.column
    if cell.value == 'Platt CPB03':
        print("Platt CPB03->" + str(cell.column))
        CPB03_Column = cell.column
    if cell.value == 'Platt CPB04':
        print("Platt CPB04->" + str(cell.column))
        CPB04_Column = cell.column
    if cell.value == 'Platt CPB05':
        print("Platt CPB05->" + str(cell.column))
        CPB05_Column = cell.column
    if cell.value == 'Platt CPB06':
        print("Platt CPB06->" + str(cell.column))
        CPB06_Column = cell.column
    if cell.value == 'Platt CPB07':
        print("Platt CPB07->" + str(cell.column))
        CPB07_Column = cell.column
    if cell.value == 'Platt CPB08':
        print("Platt CPB08->" + str(cell.column))
        CPB08_Column = cell.column
    if cell.value == 'Platt CPB09':
        print("Platt CPB09->" + str(cell.column))
        CP09_Column = cell.column
    if cell.value == 'Platt CPB10':
        print("Platt CPB10->" + str(cell.column))
        CPB10_Column = cell.column
    #Last column to search for before breaking
        break

#driver = webdriver.Chrome()
#create save ticker for save intervals
save_counter = 0

# begin main loop getting data
for x in range(2, 1000): #start no lower than 2, since 1 is the column header

    #Create info ticker for user messaging 
    New_CPB_info = 0

    # Get Source webURLlink to product data
    Src = worksheet.cell(row=x, column=Src_Column).value

    # Only continue code if there is a Src
    if Src:
        # Get Commodity Sheet Description
        ItemDescription = worksheet.cell(row=x, column=1).value

        # Get Page info
        page = requests.get(Src)

        # Create a BeautifulSoup object
        soup = BeautifulSoup(page.content, 'html.parser')


        # Get 'price' vi beautifulsoup
        try:
            Item_Cost = soup.find(class_='text--black').text
            #print (Item_Cost)    
            worksheet.cell(row=x, column=Cost_column, value=Item_Cost)
        except AttributeError:
            print ("Couldn't get cost for this line")


        #Platt product no
        if Item_Cost:
            # find the span element by class name
            span_elem = soup.find('span', {'class': 'text--secondary'})

            # get the text content of the span element
            Platt_discount_Id = span_elem.text.strip()
            Platt_discount_Id =Platt_discount_Id.replace ("Item #:","")
            Platt_discount_Id =Platt_discount_Id.replace (" ","")

            if worksheet.cell(row=x, column=ID_Column).value != Platt_discount_Id:
                worksheet.cell(row=x, column=ID_Column, value = Platt_discount_Id)

        #product UPC no
        if Item_Cost:
            # find the span element by class name
            span_elems = soup.find_all('span', {'class': 'text--secondary'})

            # get the text content of the span element
            Item_UPC =  span_elems[2].text.strip()
            Item_UPC =Item_UPC.replace ("UPC:","")
            Item_UPC =Item_UPC.replace (" ","")

            if worksheet.cell(row=x, column=Product_UPC_Column).value != Item_UPC:
                worksheet.cell(row=x, column=Product_UPC_Column, value = Item_UPC)



        
        # Write the updated time to Comodity Sheet
        worksheet.cell(row=x, column=Update_Column, value= datetime.datetime.now())


          #Give User Feedback on what scraped
        print("Platt   row:" + str(x) + " Item:" + str(ItemDescription)  + "                  $" + str(Item_Cost))

        #Make progress saves at intervals
        save_counter = save_counter + 1
        if save_counter > 100:
            wb.save(filename=workbook_loc)
            save_counter = 0
            print ("Saved")

#Save when complete
wb.save(filename=workbook_loc)
print ("Complete & Saved")
