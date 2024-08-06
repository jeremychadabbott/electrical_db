#-------------------------------------------------------------------------------------------------
#   Required Installations
#   Install BS4 package
#   Install Requests Package
#   Install Selenium
#-------------------------------------------------------------------------------------------------

import requests
import time
#Import Beautiful Soup for web scraping
from bs4 import BeautifulSoup
import pathlib
# Import openpyxl for excel workbook control
import openpyxl
from openpyxl import Workbook

# Import Selenium driver for web scraping and error handling during webscrape
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.common.exceptions import ElementNotInteractableException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
options = Options()
options.headless = True

from random import randrange                
import datetime

#Import text analysis tool
import re


# Load Workbook
filename = "Commodity Price Table"
#           Commodity Price Table
workbook_loc = str(pathlib.Path(__file__).parent.absolute()) + "\Commodity_Price_Table.xlsx" #Universal
wb = openpyxl.load_workbook(workbook_loc)

# grab the commodity table worksheet
worksheet = wb["Commodity Table"]

# iterate over the cells in the row to get excel sheet column positions 
for cell in worksheet[1]: #only scan first row
    # check if the cell's value matches the search value
        #Last column to search for before breaking
    if cell.value == 'Product_Description':
        print("Description column->" + str(cell.column))
        Description_Column = cell.column
    if cell.value == 'UPC':
        print("UPC column->" + str(cell.column))
        UPC_Column = cell.column
    if cell.value == 'Platt_UPC': #For refernce only
        print("Platt_UPC column->" + str(cell.column))
        Platt_UPC_Column = cell.column
    if cell.value == 'NorthCoastDECOdiscount_ID':
        print("ID column->" + str(cell.column))
        NorthCoastDECOdiscount_ID_Column = cell.column
    if cell.value == 'NorthCoastDECOdiscount_UPC':
        print("NorthCoastDECOdiscount UPC column->" + str(cell.column))
        Product_UPC_Column = cell.column
    if cell.value == 'NorthCoastDECOdiscount_Product':
        print("NorthCoastDECOdiscount Product name column->" + str(cell.column))
        Product_Column = cell.column
    if cell.value == 'NorthCoastDECOdiscount_Cost':
        print("Cost column->" + str(cell.column))
        Cost_column = cell.column
    if cell.value == 'NorthCoastDECOdiscount_Unit': #Ref->NorthCoastDECOdiscount
        print("Unit column->" + str(cell.column))
        Unit_Column = cell.column
    if cell.value == 'NorthCoastDECOdiscount_Src':
        print("Src column->" + str(cell.column))
        Src_Column = cell.column
    if cell.value == 'NorthCoastDECOdiscount_Update':
        print("Update column->" + str(cell.column))
        Update_Column = cell.column

    #Get Common Part Branch (CPB) Columns
    if cell.value == 'NorthCoastDECOdiscount CPB01':
        print("NorthCoastDECOdiscount CPB01->" + str(cell.column))
        CPB01_Column = cell.column
    if cell.value == 'NorthCoastDECOdiscount CPB02':
        print("NorthCoastDECOdiscount CPB02->" + str(cell.column))
        CPB02_Column = cell.column
    if cell.value == 'NorthCoastDECOdiscount CPB03':
        print("NorthCoastDECOdiscount CPB03->" + str(cell.column))
        CPB03_Column = cell.column
    if cell.value == 'NorthCoastDECOdiscount CPB04':
        print("NorthCoastDECOdiscount CPB04->" + str(cell.column))
        CPB04_Column = cell.column
    if cell.value == 'NorthCoastDECOdiscount CPB05':
        print("NorthCoastDECOdiscount CPB05->" + str(cell.column))
        CPB05_Column = cell.column
    if cell.value == 'NorthCoastDECOdiscount CPB06':
        print("NorthCoastDECOdiscount CPB06->" + str(cell.column))
        CPB06_Column = cell.column
    if cell.value == 'NorthCoastDECOdiscount CPB07':
        print("NorthCoastDECOdiscount CPB07->" + str(cell.column))
        CPB07_Column = cell.column
    if cell.value == 'NorthCoastDECOdiscount CPB08':
        print("NorthCoastDECOdiscount CPB08->" + str(cell.column))
        CPB08_Column = cell.column
    if cell.value == 'NorthCoastDECOdiscount CPB09':
        print("NorthCoastDECOdiscount CPB09->" + str(cell.column))
        CP09_Column = cell.column
    if cell.value == 'NorthCoastDECOdiscount CPB10':
        print("NorthCoastDECOdiscount CPB10->" + str(cell.column))
        CPB10_Column = cell.column
        # exit the loop once the first matching cell is found
        break

driver = webdriver.Chrome()

# Load North Coast Log In Page
driver.get('https://www.NorthCoast.com/account/login')
driver.implicitly_wait(10)

# enter the login details
username_field = driver.find_element_by_xpath('//*[@id="loginForm"]/form/div[1]/input')
username_field.send_keys('Abbott@Duttonelectric.com')

password_field = driver.find_element_by_xpath('//*[@id="loginForm"]/form/div[2]/input')
password_field.send_keys('J@bbott16')

# submit the login form
submit_button = driver.find_element_by_xpath('//*[@id="loginForm"]/form/div[4]/input')
submit_button.click()

#wait for user to navigate CAPTCHA
time.sleep(30)

# begin main loop getting data
for x in range(2, 5): #start no lower than 2, since 1 is the column header
    #create save ticker for save intervals
    save_counter = 0
    #Create info ticker for user messaging 
    New_CPB_info = 0

    # Get General Product Description 
    product_description = worksheet.cell(row=x, column=Product_Column).value

    # Get Source webURLlink to product data
    Src = worksheet.cell(row=x, column=Src_Column).value
    print (Src)
    #Test CODE >>>>> If 'none' in Source column, this is known not to be found on site, skip this item
          
    # Check if a url link to product was on the Commodity sheet, or search
    if Src:
        # A url link to product is already on the Db
        url = Src
        driver.get(url)    

        # Get 'price'
        elements = driver.find_elements_by_class_name("price-format__large-symbols")
        # iterate over the elements and print all attributes
        for element in elements:
            attributes = element.get_attribute("outerHTML")
            print(attributes)
        Item_Cost = driver.find_elements_by_class_name('price-format__large-symbols')
        print (Item_Cost)

        time.sleep(1000)


        #worksheet.cell(row=x, column=Cost_column, value=Item_Cost)

        # Write the updated time to Comodity Sheet
        worksheet.cell(row=x, column=Update_Column, value= datetime.datetime.now())

        # Get Product UPC if none exists
        if ProductUPC:
            time.sleep(.01)
        else:
            example_class_elements = soup.find_all(class_='text--secondary')
            # get the second iteration of the class
            ProductUPC = example_class_elements[3].text #Get 4th iteration
            ProductUPC = re.sub('[UPC: ]','',ProductUPC)
            ProductUPC = re.sub(r'\n', '', ProductUPC)

        # Get Product name (if none exists)
        #if ProductName:
        #    time.sleep(0.1)
        #else:    
            #ProductName = soup.find(class_='text-h3 py-1').text
            #ProductName = re.sub(r'\n', '', ProductName)
            #worksheet.cell(row=x, column=Product_Column, value=ProductName)  
        
        # Write the product ID to Commodity Sheet (if not already there)
        #if ProductID:
        #    worksheet.cell(row=x, column=ID_Column, value=ProductID)     
        #else:
        #    time.sleep(.01)

        # Get Abbott's Common part branch info (if none exists)
        if CPB02:
            time.sleep(.01)
        else:
            New_CPB_info = 1
            driver = webdriver.Chrome(chrome_options=options)
            driver.get(url)
            driver.implicitly_wait(10) # seconds
            # Get Abbott's Common Part Branch Info CPB02
            worksheet.cell(row=x, column=CPB01_Column, value="Electrical")
            try:
                CPB02 = driver.find_element("xpath","/html/body/div[1]/div/div/div/main/div/div/div[1]/div[1]/ul/a[2]").text
                worksheet.cell(row=x, column=CPB02_Column, value=CPB02)
                #print (CPB02)
            except NoSuchElementException:
                time.sleep(0.1)

            try:
            # Get Abbott's Common Part Branch Info CPB03
                CPB03 = driver.find_element("xpath","/html/body/div[1]/div/div/div/main/div/div/div[1]/div[1]/ul/a[3]").text
                worksheet.cell(row=x, column=CPB03_Column, value=CPB03)
                #print (CPB03)
            except NoSuchElementException:
                time.sleep(0.1)

            try:
                # Get Abbott's Common Part Branch Info CPB04
                CPB04 = driver.find_element("xpath","/html/body/div[1]/div/div/div/main/div/div/div[1]/div[1]/ul/a[4]").text
                worksheet.cell(row=x, column=CPB04_Column, value=CPB04)
                #print (CPB04)
            except NoSuchElementException:
                time.sleep(0.1)

            try:
                # Get Abbott's Common Part Branch Info CPB05
                CPB05 = driver.find_element("xpath","/html/body/div[1]/div/div/div/main/div/div/div[1]/div[1]/ul/a[5]").text
                worksheet.cell(row=x, column=CPB05_Column, value=CPB05)
                #print (CPB05)
            except NoSuchElementException:
                time.sleep(0.1)

        #Give User Feedback on what scraped
        print("NorthCoastDECOdiscount   row:" + str(x) + " Item:" + str(ProductName)  + "                  $" + str(Item_Cost))
        if New_CPB_info  == 1:
            print("NorthCoastDECOdiscount Tree-> Electrical->" + str(CPB02) + "->" + str(CPB03) + "->" + str(CPB04) + "->" + str(CPB05))

        # Write the product UPC to Commodity Sheet (if not already there)
        if ProductUPC:
            worksheet.cell(row=x, column=Product_UPC_Column, value=ProductUPC)     
        else:
            time.sleep(.01) 

        #Make progress saves at intervals
        save_counter = save_counter + 1
        if save_counter > 100:
            wb.save(filename=workbook_loc)
            save_counter = 0
            print ("Saved")

#Save when complete
wb.save(filename=workbook_loc)
print ("Complete & Saved")
