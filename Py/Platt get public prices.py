#-------------------------------------------------------------------------------------------------
#   Required Installations
#   Install BS4 package
#   Install Requests Package
#   Install Selenium
#-------------------------------------------------------------------------------------------------


import requests
import time
#Import Beautiful Soup for web scraping
from bs4 import BeautifulSoup
import pathlib
# Import openpyxl for excel workbook control
import openpyxl
from openpyxl import Workbook

# Import Selenium driver for web scraping and error handling during webscrape
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.common.exceptions import ElementNotInteractableException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
options = Options()
options.headless = True
driver = webdriver.Chrome(chrome_options=options)


from random import randrange                
import datetime

#Import text analysis tool
import re


# Load Workbook
filename = "Commodity Price Table"
#           Commodity Price Table
workbook_loc = str(pathlib.Path(__file__).parent.absolute()) + "\Commodity_Price_Table.xlsx" #Universal
wb = openpyxl.load_workbook(workbook_loc)

# grab the commodity table worksheet
worksheet = wb["Commodity Table"]

# iterate over the cells in the row to get excel sheet column positions 
for cell in worksheet[1]: #only scan first row
    # check if the cell's value matches the search value
    if cell.value == 'UPC':
        print("UPC column->" + str(cell.column))
        UPC_Column = cell.column
    if cell.value == 'Platt_ID':
        print("ID column->" + str(cell.column))
        ID_Column = cell.column
    if cell.value == 'Platt_Product':
        print("Description column->" + str(cell.column))
        Description_Column = cell.column
    if cell.value == 'Platt_UPC':
        print("Platt UPC column->" + str(cell.column))
        Product_UPC_Column = cell.column
    if cell.value == 'Platt_Product':
        print("Platt Product name column->" + str(cell.column))
        Product_Column = cell.column
    if cell.value == 'Platt_Cost':
        print("Cost column->" + str(cell.column))
        Cost_column = cell.column
    if cell.value == 'Platt_Unit':
        print("Unit column->" + str(cell.column))
        Unit_Column = cell.column
    if cell.value == 'Platt_Src':
        print("Src column->" + str(cell.column))
        Src_Column = cell.column
    if cell.value == 'Platt_Update':
        print("Update column->" + str(cell.column))
        Update_Column = cell.column

    #Get Common Part Branch (CPB) Columns
    if cell.value == 'Platt CPB01':
        print("Platt CPB01->" + str(cell.column))
        CPB01_Column = cell.column
    if cell.value == 'Platt CPB02':
        print("Platt CPB02->" + str(cell.column))
        CPB02_Column = cell.column
    if cell.value == 'Platt CPB03':
        print("Platt CPB03->" + str(cell.column))
        CPB03_Column = cell.column
    if cell.value == 'Platt CPB04':
        print("Platt CPB04->" + str(cell.column))
        CPB04_Column = cell.column
    if cell.value == 'Platt CPB05':
        print("Platt CPB05->" + str(cell.column))
        CPB05_Column = cell.column
    if cell.value == 'Platt CPB06':
        print("Platt CPB06->" + str(cell.column))
        CPB06_Column = cell.column
    if cell.value == 'Platt CPB07':
        print("Platt CPB07->" + str(cell.column))
        CPB07_Column = cell.column
    if cell.value == 'Platt CPB08':
        print("Platt CPB08->" + str(cell.column))
        CPB08_Column = cell.column
    if cell.value == 'Platt CPB09':
        print("Platt CPB09->" + str(cell.column))
        CP09_Column = cell.column
    if cell.value == 'Platt CPB10':
        print("Platt CPB10->" + str(cell.column))
        CPB10_Column = cell.column
    #Last column to search for before breaking
        break

driver = webdriver.Chrome()

# begin main loop getting data
for x in range(2, 580): #start no lower than 2, since 1 is the column header
    #create save ticker for save intervals
    save_counter = 0
    #Create info ticker for user messaging 
    New_CPB_info = 0

    # Get Source webURLlink to product data
    Src = worksheet.cell(row=x, column=Src_Column).value

    # Only continue code if there is a Src
    if Src:
    

        # Use the folowing code to screen for "search results pages" in the URL
        if Src:
            if Src.find('search') == -1:
                time.sleep(0.1)
            else:
                Src = ""

        # Get Commodity Sheet Description
        ItemDescription = worksheet.cell(row=x, column=Description_Column).value
        # Get Commodity Sheet Description VendorProductID (Per Vendor)
        VendorProductID = worksheet.cell(row=x, column=ID_Column).value
        # ProductUPC Commodity Sheet Description (Per manufacturer of item)
        ProductUPC = worksheet.cell(row=x, column=Product_UPC_Column).value
        # Product Name
        ProductName = worksheet.cell(row=x, column =Product_Column).value
        # Get Common Parts Branch Descriptions
        CPB02 = worksheet.cell(row=x, column =CPB02_Column).value
        CPB03 = worksheet.cell(row=x, column =CPB03_Column).value
        CPB04 = worksheet.cell(row=x, column =CPB04_Column).value
        CPB05 = worksheet.cell(row=x, column =CPB05_Column).value

        # Check for Produxt URL link first
        if Src:
            # A url link to product is already on the Db
            url = Src
        else:
            # There's no url link on commodity sheet, so use website search function
            url = "https://www.platt.com/s/search?q=" + str(ItemDescription)
            driver.get(url)
            print("")
            print("Used Search to find: " + str(ItemDescription) + "            ")
            print ("")
            #Wait to load
            driver.implicitly_wait(200)
        

            # Ask for user input to set the correct url link to catalog page. This pause for user input allows user to navigate chrome if needed.
            string = input("Navigate to correct page and hit <enter>")
            if string:
                url = driver.current_url
                print(string)
            else:
                #Write no results to Worksheet
                url = driver.current_url
                print(string)
            driver.close ()

        # Use Beautiful Soup
        # Get Page info
        page = requests.get(url)

        # Create a BeautifulSoup object
        soup = BeautifulSoup(page.content, 'html.parser')

        # Get Product URLLink 
        if Src:
            time.sleep(.01)
        else:
            worksheet.cell(row=x, column=Src_Column, value=url)
            wb.save(filename=workbook_loc)
            save_counter = 0
            print ("Saved")  

        # Get Product name (if none exists)

        # Get 'price' vi beautifulsoup
        Item_Cost = soup.find(class_='text--black').text

        # Anomalies / Price modifiers
        #if Src.ProductName('search') == -1:
        #    time.sleep(0.1)
        #else:
        #    Src = ""

        # Write Price to commodities Table
        #print (Item_Cost)    
        worksheet.cell(row=x, column=Cost_column, value=Item_Cost)
        
        
        # Write the updated time to Comodity Sheet
        worksheet.cell(row=x, column=Update_Column, value= datetime.datetime.now())

        # Get Product UPC if none exists
        if ProductUPC:
            time.sleep(.01)
        else:
            example_class_elements = soup.find_all(class_='text--secondary')
            # get the second iteration of the class
            ProductUPC = example_class_elements[3].text #Get 4th iteration
            ProductUPC = re.sub('[UPC: ]','',ProductUPC)
            ProductUPC = re.sub(r'\n', '', ProductUPC)
        
        # Write the product ID to Commodity Sheet (if not already there)
        #if ProductID:
        #    worksheet.cell(row=x, column=ID_Column, value=ProductID)     
        #else:
        #    time.sleep(.01)

        # Get Abbott's Common part branch info (if none exists)
        if CPB02:
            time.sleep(.01)
        else:
            New_CPB_info = 1
            driver = webdriver.Chrome(chrome_options=options)
            driver.get(url)
            driver.implicitly_wait(10) # seconds
            # Get Abbott's Common Part Branch Info CPB02
            worksheet.cell(row=x, column=CPB01_Column, value="Electrical")
            try:
                CPB02 = driver.find_element("xpath","/html/body/div[1]/div/div/div/main/div/div/div[1]/div[1]/ul/a[2]").text
                worksheet.cell(row=x, column=CPB02_Column, value=CPB02)
                #print (CPB02)
            except NoSuchElementException:
                time.sleep(0.1)

            try:
            # Get Abbott's Common Part Branch Info CPB03
                CPB03 = driver.find_element("xpath","/html/body/div[1]/div/div/div/main/div/div/div[1]/div[1]/ul/a[3]").text
                worksheet.cell(row=x, column=CPB03_Column, value=CPB03)
                #print (CPB03)
            except NoSuchElementException:
                time.sleep(0.1)

            try:
                # Get Abbott's Common Part Branch Info CPB04
                CPB04 = driver.find_element("xpath","/html/body/div[1]/div/div/div/main/div/div/div[1]/div[1]/ul/a[4]").text
                worksheet.cell(row=x, column=CPB04_Column, value=CPB04)
                #print (CPB04)
            except NoSuchElementException:
                time.sleep(0.1)

            try:
                # Get Abbott's Common Part Branch Info CPB05
                CPB05 = driver.find_element("xpath","/html/body/div[1]/div/div/div/main/div/div/div[1]/div[1]/ul/a[5]").text
                worksheet.cell(row=x, column=CPB05_Column, value=CPB05)
                #print (CPB05)
            except NoSuchElementException:
                time.sleep(0.1)

        #Give User Feedback on what scraped
        print("Platt   row:" + str(x) + " Item:" + str(ProductName)  + "                  $" + str(Item_Cost))
        if New_CPB_info  == 1:
            print("Platt Tree-> Electrical->" + str(CPB02) + "->" + str(CPB03) + "->" + str(CPB04) + "->" + str(CPB05))

        # Write the product UPC to Commodity Sheet (if not already there)
        if ProductUPC:
            worksheet.cell(row=x, column=Product_UPC_Column, value=ProductUPC)     
        else:
            time.sleep(.01) 

        #Make progress saves at intervals
        save_counter = save_counter + 1
        if save_counter > 100:
            wb.save(filename=workbook_loc)
            save_counter = 0
            print ("Saved")

#Save when complete
wb.save(filename=workbook_loc)
print ("Complete & Saved")
