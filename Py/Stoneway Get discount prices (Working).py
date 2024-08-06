




#-------------------------------------------------------------------------------------------------
#   Required Installations
#   Install BS4 package
#   Install Requests Package
#   Install Selenium
#-------------------------------------------------------------------------------------------------

import requests
import time
#Import Beautiful Soup for web scraping
from bs4 import BeautifulSoup
import pathlib
# Import openpyxl for excel workbook control
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import numbers

# Import Selenium driver for web scraping and error handling during webscrape
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.common.exceptions import ElementNotInteractableException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
options = Options()
options.headless = True

from random import randrange                
import datetime

#Import text analysis tool
import re


# Load Workbook
filename = "Commodity Price Table"
#           Commodity Price Table
workbook_loc = str(pathlib.Path(__file__).parent.parent.absolute()) + "\Commodity_Price_Table.xlsx" #Universal
wb = openpyxl.load_workbook(workbook_loc)

# grab the commodity table worksheet
worksheet = wb["Commodity Table"]

# iterate over the cells in the row to get excel sheet column positions 
for cell in worksheet[1]: #only scan first row
    # check if the cell's value matches the search value
        #Last column to search for before breaking
    if cell.value == 'Product_Description':
        print("Description column->" + str(cell.column))
        Description_Column = cell.column
    if cell.value == 'UPC':
        print("UPC column->" + str(cell.column))
        UPC_Column = cell.column
    if cell.value == 'Platt_UPC': #For refernce only
        print("Platt_UPC column->" + str(cell.column))
        Platt_UPC_Column = cell.column
    if cell.value == 'Stoneway_discount_ID':
        print("ID column->" + str(cell.column))
        Stoneway_discount_ID_Column = cell.column
    if cell.value == 'Stoneway_discount_UPC':
        print("Stoneway_discount UPC column->" + str(cell.column))
        Product_UPC_Column = cell.column
    if cell.value == 'Stoneway_discount_Product':
        print("Stoneway_discount Product name column->" + str(cell.column))
        Product_Column = cell.column
    if cell.value == 'Stoneway_discount_Cost':
        print("Cost column->" + str(cell.column))
        Cost_column = cell.column
    if cell.value == 'Stoneway_discount_Unit': #Ref->Stoneway_discount
        print("Unit column->" + str(cell.column))
        Unit_Column = cell.column
    if cell.value == 'Stoneway_discount_Src':
        print("Src column->" + str(cell.column))
        Src_Column = cell.column
    if cell.value == 'Stoneway_discount_Update':
        print("Update column->" + str(cell.column))
        Update_Column = cell.column
        break


driver = webdriver.Chrome()
#driver = webdriver.Chrome(chrome_options=options)

# Load Stoneway Log In Page
driver.get('https://www.stoneway.com/index.cfm?dsp=member.login')
driver.implicitly_wait(10)

# enter the login details
username_field = driver.find_element ('xpath','//*[@id="user_email"]')
username_field.send_keys('Abbott@Duttonelectric.com')

password_field = driver.find_element ('xpath','//*[@id="user_pass"]')
password_field.send_keys('J@bbott16')

# submit the login form
submit_button = driver.find_element ('xpath', '//*[@id="signin_block"]/input')
submit_button.click()
driver.implicitly_wait(10)

#wait for user to navigate CAPTCHA
#time.sleep(30)

# begin main loop getting data
for x in range(2, 1000): #start no lower than 2, since 1 is the column header
    #create save ticker for save intervals
    save_counter = 0

    # Get General Product Description 
    product_description = worksheet.cell(row=x, column=1).value

    # Get Source webURLlink to product data
    Src = worksheet.cell(row=x, column=Src_Column).value
  
          
    # Check if a url link to product was on the Commodity sheet, or search
    if Src:
        if ".stoneway.c" in Src:
            if ".search" not in Src:
                # A url link to product is already on the Db
                driver.get(Src)    
                # Create a BeautifulSoup object
                html = driver.page_source
                soup = BeautifulSoup(html, "html.parser")
                # Get 'price' vi beautifulsoup
                priceperqty = soup.find(class_='price_corresponding').text

                #priceperqty = driver.find_element_by_class_name("price_corresponding").text
                priceperqty = priceperqty.replace("$","")
                #print (priceperqty)

                uom = soup.find("b", class_="price_per").text
                if "each" not in uom:
                    time.sleep(0.1)
                #print (uom)

                # Get Stoneway product ID
                if priceperqty != 0:
                    # find the span element by class name
                    span_elems = soup.find_all('td')

                    # get the text content of the span element
                    Stoneway_discount_ID = span_elems[2].text
            
                    # get stoneway UPC
                    Stoneway_discount_UPC = span_elems[3].text

                    #WRite info to Db
                    if worksheet.cell(row=x, column=Stoneway_discount_ID_Column).value != Stoneway_discount_ID:
                        worksheet.cell(row=x, column=Stoneway_discount_ID_Column).number_format = numbers.FORMAT_TEXT
                        worksheet.cell(row=x, column=Stoneway_discount_ID_Column, value = Stoneway_discount_ID)
                        
                    if worksheet.cell(row=x, column=Product_UPC_Column).value != Stoneway_discount_UPC:
                        worksheet.cell(row=x, column=Product_UPC_Column).number_format = numbers.FORMAT_TEXT
                        worksheet.cell(row=x, column=Product_UPC_Column, value = Stoneway_discount_UPC)

                #sku_element = driver.find_element_by_id("sku_numb").text
                sku_element = soup.find("span", id="sku_numb")

                # Extract the text from the element
                sku_text = sku_element.text


                # write price to the worksheet
                worksheet.cell(row=x, column=Cost_column, value=priceperqty)

                # Write the updated time to Comodity Sheet
                worksheet.cell(row=x, column=Update_Column, value= datetime.datetime.now())


                #Give User Feedback on what scraped
                print("Stoneway_discount   row:" + str(x) + " " + str(sku_element) +     " $" + str(priceperqty) + " " + str(uom))
        

                #Make progress saves at intervals
                save_counter = save_counter + 1
                if save_counter > 100:
                    wb.save(filename=workbook_loc)
                    save_counter = 0
                    print ("Saved")

#Save when complete
wb.save(filename=workbook_loc)
print ("Complete & Saved")
