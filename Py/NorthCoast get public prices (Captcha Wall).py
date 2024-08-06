#-------------------------------------------------------------------------------------------------
#   Required Installations
#   Install BS4 package
#   Install Requests Package
#   Install Selenium
#-------------------------------------------------------------------------------------------------

import requests
import time
#Import Beautiful Soup for web scraping
from bs4 import BeautifulSoup
import pathlib
# Import openpyxl for excel workbook control
import openpyxl
from openpyxl import Workbook

# Import Selenium driver for web scraping and error handling during webscrape
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.common.exceptions import ElementNotInteractableException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
options = Options()
options.headless = True

from random import randrange                
import datetime

#Import text analysis tool
import re


# Load Workbook
filename = "Commodity Price Table"
#           Commodity Price Table
workbook_loc = str(pathlib.Path(__file__).parent.parent.absolute()) + "\Commodity_Price_Table.xlsx" #Universal
wb = openpyxl.load_workbook(workbook_loc)

# grab the commodity table worksheet
worksheet = wb["Commodity Table"]

# iterate over the cells in the row to get excel sheet column positions 
for cell in worksheet[1]: #only scan first row
    # check if the cell's value matches the search value
        #Last column to search for before breaking
    if cell.value == 'NorthCoastdiscount_ID':
        print("ID column->" + str(cell.column))
        NorthCoastdiscount_ID_Column = cell.column
    if cell.value == 'NorthCoastdiscount_UPC':
        print("NorthCoastdiscount UPC column->" + str(cell.column))
        Product_UPC_Column = cell.column
    if cell.value == 'NorthCoastdiscount_Product':
        print("NorthCoastdiscount Product name column->" + str(cell.column))
        Product_Column = cell.column
    if cell.value == 'NorthCoastdiscount_Cost':
        print("Cost column->" + str(cell.column))
        Cost_column = cell.column
    if cell.value == 'NorthCoastdiscount_Unit': #Ref->NorthCoastdiscount
        print("Unit column->" + str(cell.column))
        Unit_Column = cell.column
    if cell.value == 'NorthCoastdiscount_Src':
        print("Src column->" + str(cell.column))
        Src_Column = cell.column
    if cell.value == 'NorthCoastdiscount_Update':
        print("Update column->" + str(cell.column))
        Update_Column = cell.column

driver = webdriver.Chrome()

# Load North Coast Log In Page
#driver.get('https://www.NorthCoast.com/account/login')
#driver.implicitly_wait(10)

# enter the login details
#username_field = driver.find_element_by_xpath('//*[@id="loginForm"]/form/div[1]/input')
#username_field.send_keys('Abbott@Duttonelectric.com')

#password_field = driver.find_element_by_xpath('//*[@id="loginForm"]/form/div[2]/input')
#password_field.send_keys('J@bbott16')

# submit the login form
#submit_button = driver.find_element_by_xpath('//*[@id="loginForm"]/form/div[4]/input')
#submit_button.click()

#wait for user to navigate CAPTCHA
#time.sleep(30)

# Set save_counter variable
save_counter = 0

# begin main loop getting data
for x in range(2, 600): #start no lower than 2, since 1 is the column header


    # Get Source webURLlink to product data
    Src = worksheet.cell(row=x, column=Src_Column).value
          
    # Check if a url link to product was on the Commodity sheet, or search
    if Src:
        # A url link to product is already on the Db
        driver.get(Src)    

        # Get 'price'
        priceperqty = driver.find_element_by_class_name("pricerPerQty").text
        priceperqty = priceperqty.replace("$","")
        priceperqty = float(priceperqty)
        #print (priceperqty)

        uom = driver.find_element_by_class_name("priceUOM").text
        if "c" in uom:
            priceperqty = priceperqty / 100

        sku_element = driver.find_element_by_tag_name("h1").text
        # Write price to worksheet
        #worksheet.cell(row=x, column=Cost_column, value=Item_Cost)

        # Write the updated time to Comodity Sheet
        worksheet.cell(row=x, column=Update_Column, value= datetime.datetime.now())

        #Give User Feedback on what scraped
        print("NorthCoast   row:"  + str(x) + " " + str(sku_element) +     " $" + str(priceperqty) + " " + str(uom))

        #Make progress saves at intervals
        save_counter = save_counter + 1
        if save_counter > 100:
            wb.save(filename=workbook_loc)
            save_counter = 0
            print ("Saved")

#Save when complete
wb.save(filename=workbook_loc)
print ("Complete & Saved")
